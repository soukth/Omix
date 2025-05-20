import PySimpleGUI as sg
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook, Workbook
import requests
from bs4 import BeautifulSoup
import threading
import queue
from tqdm import tqdm
import random
import openpyxl
from openpyxl.styles import Font, Alignment
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from matplotlib_venn import venn2, venn3


class ProteinFinder:
    def __init__(self):
        self.data = None
        self.file_path = None
        self.transporter_proteins = ["ABCB1", "MDR1", "ABCB4", "MDR4", "ABCC1", "ABCC2", "ABCC3", "ABCC4", "ABCC5", "ABCC6", "ABCC10", "ABCC11", "ABCG2",
                                     "SLC6A4", "SLC6A2", "SLC22A1", "SLC22A2", "SLC22A3", "SLC22A4", "SLC22A5", "SLC22A6", "SLC22A7", "SLC22A8",
                                     "SLC22A9", "SLC21", "SLCO"]
        self.progesterone_proteins = ["PRA", "PR-A", "PRB", "PR-B", "mPR", "PGR", "PGRMC1", "PGRMC2", "PAQR7", "PAQR8", "PAQR5", "PAQR6", "PAQR9"]
        self.prolactin_proteins = ["PRL"]  # Will be populated dynamically
        hdac_proteins = [f"HDAC{i}" for i in range(1, 12)]  # HDAC1 to HDAC11
        med_proteins = [f"MED{i}" for i in range(1, 20)]  # MED1 to MED20
        self.transcription_factors = ["SWI", "SNF", "AP-1", "AP1", "SP1", "CARM1"] + hdac_proteins + med_proteins
        self.window_stack = []

    def assign_to_group(self, citation, source):
        """
        Dynamically assign the protein to a functional group based on citation content from a specific AI search engine (PubMed, NCBI, Scopus).
        """
        # First, check the source of the citation
        if source == 'PubMed':
            return self.assign_pubmed_group(citation)
        elif source == 'NCBI':
            return self.assign_ncbi_group(citation)
        elif source == 'Scopus':
            return self.assign_scopus_group(citation)
        else:
            return "Unclassified"  # If source is unknown

    def assign_pubmed_group(self, citation):
        """
        Logic to assign a group based on PubMed citation text
        """
        if "cancer" in citation.lower():
            return "Cancer-related Proteins"
        elif "neurotransmitter" in citation.lower():
            return "Neurotransmitter Transporters"
        elif "inflammation" in citation.lower():
            return "Inflammatory Proteins"
        elif "apoptosis" in citation.lower():
            return "Apoptosis-related Proteins"
        else:
            return "PubMed-Related Proteins"

    def assign_ncbi_group(self, citation):
        """
        Logic to assign a group based on NCBI citation text
        """
        if "metabolism" in citation.lower():
            return "Metabolic Pathways Proteins"
        elif "immune" in citation.lower():
            return "Immune Response Proteins"
        elif "DNA repair" in citation.lower():
            return "DNA Repair Proteins"
        else:
            return "NCBI-Related Proteins"

    def assign_scopus_group(self, citation):
        """
        Logic to assign a group based on Scopus citation text
        """
        if "signaling" in citation.lower():
            return "Signaling Pathways Proteins"
        elif "receptor" in citation.lower():
            return "Receptor Proteins"
        elif "cardiovascular" in citation.lower():
            return "Cardiovascular-related Proteins"
        else:
            return "Scopus-Related Proteins"

    def upload_file(self):
        file_path = sg.popup_get_file('Select a file')
        if file_path:
            self.data = pd.read_excel(file_path) if file_path.endswith(".xlsx") else pd.read_csv(file_path)
            self.file_path = file_path
            sg.popup("File uploaded successfully")

    def normalize_data(self):
        if self.data is not None:
            # Convert all columns to numeric, coerce errors to NaN, and fill NaN with 0
            self.data = self.data.apply(pd.to_numeric, errors='coerce').fillna(0)
            # Normalize the data
            self.data = (self.data - self.data.min()) / (self.data.max() - self.data.min())

    def show_heatmap(self):
        if self.data is not None:
            # Check if the data is properly arranged
            if len(self.data.columns) < 3:
                sg.popup_error("Data not arranged properly. At least 3 columns are required.")
                return

            # Use all columns except the first one (assuming first column is gene names/IDs)
            data_subset = self.data.iloc[:, 1:]
            
            # Check if the data can be converted to numeric
            try:
                # Convert to numeric and handle any non-numeric values
                data_subset = data_subset.apply(pd.to_numeric, errors='coerce')
                
                # Check if we have any valid numeric data left
                if data_subset.isnull().all().all():
                    sg.popup_error("No valid numeric data found. Please check your input file.")
                    return
                
                # Fill NaN with 0 and normalize the data
                data_subset = data_subset.fillna(0)
                data_subset = (data_subset - data_subset.min()) / (data_subset.max() - data_subset.min())
                
                # Create the heatmap
                fig = px.imshow(data_subset.T,  # Transpose for better visualization
                                labels=dict(x="Genes", y="Conditions", color="Expression"),
                                x=self.data.iloc[:, 0],  # Use the first column for gene names
                                y=data_subset.columns,
                                aspect="auto",
                                color_continuous_scale="RdBu_r")
                
                # Update layout
                fig.update_layout(
                    title="Gene Expression Heatmap",
                    xaxis_title="Genes",
                    yaxis_title="Conditions",
                    width=1000,
                    height=800,
                    margin=dict(l=50, r=50, b=100, t=100, pad=4)
                )
                
                # Show the plot
                fig.show()

                # Save the heatmap as an HTML file
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                heatmap_path = os.path.join(desktop, "interactive_heatmap.html")
                fig.write_html(heatmap_path)
                sg.popup(f"Interactive heatmap saved to {heatmap_path}")
            
            except Exception as e:
                sg.popup_error(f"Error processing data: {str(e)}")
        else:
            sg.popup_error("No data loaded. Please upload a file first.")

    def get_pubmed_citation(self, protein_name):
        base_url = "https://pubmed.ncbi.nlm.nih.gov"
        search_url = f"{base_url}/?term={protein_name}"

        try:
            response = requests.get(search_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, "html.parser")
            article_link = soup.find('a', {'class': 'docsum-title'})['href']
            citation_url = base_url + article_link

            return citation_url
        except Exception as e:
            print(f"Error retrieving PubMed citation for {protein_name}: {e}")
            return "PubMed citation not found"

    def get_ncbi_citation(self, protein_name):
        base_url = "https://www.ncbi.nlm.nih.gov"
        search_url = f"{base_url}/nuccore/?term={protein_name}"

        try:
            response = requests.get(search_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, "html.parser")
            article_link = soup.find('a', {'class': 'title'})['href']
            citation_url = base_url + article_link

            return citation_url
        except Exception as e:
            print(f"Error retrieving NCBI citation for {protein_name}: {e}")
            return "NCBI citation not found"

    def get_scopus_citation(self, protein_name):
        base_url = "https://www.scopus.com"
        search_url = f"{base_url}/results/results.uri?sort=plf-f&src=s&sid=7b9da35a80020179a033c91a05380ac1&sot=b&sdt=b&sl=63&q={protein_name}"

        try:
            response = requests.get(search_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, "html.parser")
            article_link = soup.find('a', {'class': 'docs-title'})['href']
            citation_url = base_url + article_link

            return citation_url
        except Exception as e:
            print(f"Error retrieving Scopus citation for {protein_name}: {e}")
            return "Scopus citation not found"

    def query_perplexity(self, gene_ids):
        """
        Simulated Perplexity AI query function
        Replace this with actual API call to Perplexity AI
        """
        # Simulate dynamic functional groups
        possible_groups = [
            "Immune Response Regulation",
            "Cell Cycle Control",
            "Apoptosis Signaling",
            "Transcriptional Regulation",
            "DNA Repair Mechanisms",
            "Protein Folding and Stability",
            "Cellular Stress Response",
            "Neurotransmitter Signaling",
            "Lipid Metabolism",
            "Extracellular Matrix Organization"
        ]
        
        # Randomly select 3 groups for this simulation
        functional_groups = random.sample(possible_groups, 3)
        
        categorized_genes = {group: [] for group in functional_groups}
        manuscripts = {}
        
        for gene in gene_ids:
            group = random.choice(functional_groups)
            categorized_genes[group].append(gene)
            
            # Simulate manuscript IDs (PMID for PubMed, PMC for PubMed Central)
            manuscripts[gene] = [f"PMID:{random.randint(10000000, 99999999)}" for _ in range(2)] + \
                                [f"PMC{random.randint(1000000, 9999999)}" for _ in range(2)]
        
        return categorized_genes, manuscripts

    def show_functional_analysis(self):
        if self.data is not None:
            genes = self.data.iloc[:, 0].tolist()
            
            # Get protein categorization and manuscripts from Perplexity AI (simulated)
            protein_groups, manuscripts = self.query_perplexity(genes)
            
            # Create new workbook
            wb = openpyxl.Workbook()
            
            # Create sheets for each functional group
            for group_name, proteins in protein_groups.items():
                ws = wb.create_sheet(title=group_name[:31])  # Excel sheet names are limited to 31 characters
                
                # Write headers
                headers = ['Gene ID', 'Manuscript IDs']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Populate data
                for row, protein in enumerate(proteins, start=2):
                    ws.cell(row=row, column=1, value=protein)
                    ws.cell(row=row, column=2, value=', '.join(manuscripts[protein]))
            
            # Remove default sheet
            wb.remove(wb['Sheet'])
            
            # Save workbook at the same location as input file using the input file name
            input_filename = os.path.splitext(os.path.basename(self.file_path))[0]
            output_file = os.path.join(os.path.dirname(self.file_path), f'{input_filename}_functional_groups.xlsx')
            wb.save(output_file)
            sg.popup(f"Categorized genes saved to {output_file}")
        
    def create_excel(self, protein_groups):
        # Create a DataFrame for the functional analysis results
        # Find the maximum length of the protein lists
        max_length = max(len(proteins) for proteins in protein_groups.values())

        # Create a dictionary for each group with protein data
        excel_data = {}
        for group, proteins in protein_groups.items():
            excel_data[group] = proteins + [''] * (max_length - len(proteins))

        # Create a DataFrame for the functional analysis results
        results_df = pd.DataFrame(excel_data)

        # Add a reference list at the bottom
        reference_list = []
        citation_number = 1
        for group, proteins in protein_groups.items():
            for protein in proteins:
                if protein:
                    reference_list.append(f"{citation_number}. {protein.split('(')[-1][:-1]}")  # Clean citation from protein string
                    citation_number += 1

        # Add the reference list to the DataFrame
        ref_column = [''] * (max_length - len(reference_list)) + reference_list
        results_df['References'] = ref_column

        # Write the updated DataFrame to Excel
        input_filename = os.path.splitext(os.path.basename(self.file_path))[0]
        output_filename = f"{input_filename}_Functional_Analysis_Results.xlsx"
        desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
        excel_path = os.path.join(desktop, output_filename)

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name="Analysis")

        # Notify the user that the Excel file has been saved
        sg.popup(f"Functional analysis results saved as {output_filename}")

    def show_protein_options(self):
        layout = [
            [sg.Button('Targeted Analysis', size=(20, 2), font='Any 12')],
            [sg.Button('Heatmap', size=(20, 2), font='Any 12')],
            [sg.Button('Volcano Plot', size=(20, 2), font='Any 12')],
            [sg.Button('Venn Diagram', size=(20, 2), font='Any 12')],
            [sg.Button('Functional Analysis', size=(20, 2), font='Any 12')],
            [sg.Button('Back', size=(20, 2), font='Any 12')]
        ]
        self.window_stack.append(sg.Window('Protein Options', layout, size=(550, 400), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event == 'Targeted Analysis':
                self.upload_file()
                self.show_targeted_analysis()
            elif event == 'Heatmap':
                self.upload_file()
                self.show_heatmap()
            elif event == 'Volcano Plot':
                self.upload_file()
                self.show_volcano_plot()
            elif event == 'Venn Diagram':
                self.show_venn_diagram()
            elif event == 'Functional Analysis':
                self.upload_file()
                self.show_functional_analysis()
            elif event == 'Back' or event == sg.WINDOW_CLOSED:
                self.window_stack[-1].close()
                self.window_stack.pop()
                return
            self.window_stack[-1].close()

    def show_targeted_analysis(self):
        button_layout = [
            [sg.Button('Cell Fate', size=(25, 2), font='Any 8'),
            sg.Button('Epithelial to Mesenchymal Markers', size=(25, 2), font='Any 7')],
            [sg.Button('Exosome', size=(25, 2), font='Any 8'),
            sg.Button('Immune Response', size=(25, 2), font='Any 8')],
            [sg.Button('Inflammation', size=(25, 2), font='Any 8'),
            sg.Button('Neurotransmitters', size=(25, 2), font='Any 8')],
            [sg.Button('Stem Cell', size=(25, 2), font='Any 8'),
            sg.Button('Toll-Like Receptor', size=(25, 2), font='Any 8')],
            [sg.Button('Transporter Proteins', size=(25, 2), font='Any 8'),
            sg.Button('Progesterone Proteins', size=(25, 2), font='Any 8')],
            [sg.Button('Prolactin Proteins', size=(25, 2), font='Any 8'),
            sg.Button('Transcription factors', size=(25, 2), font='Any 8')],
            [sg.Button('Back', size=(25, 2), font='Any 8', expand_x=True)]
        ]

        layout = [
            [sg.Text('Targeted Analysis', font='Any 16', justification='center', expand_x=True)],
            [sg.Column(button_layout, element_justification='center')]
        ]

        self.window_stack.append(sg.Window('Targeted Analysis', layout, size=(600, 450), element_justification='center'))
        
        while True:
            event, values = self.window_stack[-1].read()
            if event == 'Cell Fate':
                self.show_cell_fate_options()
            elif event == 'Epithelial to Mesenchymal Markers':
                self.show_emt_options()
            elif event == 'Exosome':
                self.show_exosome_options()
            elif event == 'Immune Response':
                self.show_immune_response_options()
            elif event == 'Inflammation':
                self.show_inflammation_options()
            elif event == 'Neurotransmitters':
                self.show_neurotransmitter_options()
            elif event == 'Stem Cell':
                self.show_stem_cell_options()
            elif event == 'Toll-Like Receptor':
                self.show_tlr_options()
            elif event == 'Transporter Proteins':
                self.show_transporter_proteins()
            elif event == 'Progesterone Proteins':
                self.show_progesterone_proteins()
            elif event == 'Prolactin Proteins':
                self.show_prolactin_proteins()
            elif event == 'Transcription factors':
                self.show_transcription_factors()
            elif event == 'Back' or event == sg.WINDOW_CLOSED:
                self.window_stack[-1].close()
                self.window_stack.pop()
                return
        self.window_stack[-1].close()

    def show_cell_fate_options(self):
        layout = [
            [sg.Button('Apoptotic Signalling Pathway', size=(22, 2)),
            sg.Button('Apoptotic', size=(22, 2))],
            [sg.Button('Autophagy', size=(22, 2)),
            sg.Button('Cellular Senescence', size=(22, 2))],
            [sg.Button('Necrotic Cell Death', size=(22, 2)),
            sg.Button('Back', size=(22, 2))]
            
        ]
        self.window_stack.append(sg.Window('Cell Fate Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()

    def show_emt_options(self):
        layout = [
            [sg.Button('Cell Adhesion', size=(23, 2)),
            sg.Button('EPT', size=(23, 2))],
            [sg.Button('Extracellular Matrix Organization', size=(23, 2)),
            sg.Button('TGF-Beta Signallisng Pathway', size=(23, 2))],
            [sg.Button('Back', size=(22, 2),expand_x=True)]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()
    def show_exosome_options(self):
        layout = [
            [sg.Button('Exosome Marker List', size=(22, 2))],
            
            [sg.Button('Back', size=(22, 2))]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()
    def show_immune_response_options(self):
        layout = [
            [sg.Button('Exosome Marker List', size=(22, 2))],
            
            [sg.Button('Back', size=(22, 2))]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()

    def show_inflammation_options(self):
        layout = [
            [sg.Button('Exosome Marker List', size=(22, 2))],
            
            [sg.Button('Back', size=(22, 2))]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()
    def show_neurotransmitter_options(self):
        layout = [
            [sg.Button('Exosome Marker List', size=(22, 2))],
            
            [sg.Button('Back', size=(22, 2))]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()
    
    def show_stem_cell_options(self):
        layout = [
            [sg.Button('Exosome Marker List', size=(22, 2))],
            
            [sg.Button('Back', size=(22, 2))]
        ]
        self.window_stack.append(sg.Window('EMT Options', layout, size=(500, 250), element_justification='center'))
        while True:
            event, values = self.window_stack[-1].read()
            if event in (sg.WINDOW_CLOSED, 'Back'):
                break
            # Add functionality for each sub-button here
        self.window_stack[-1].close()
        self.window_stack.pop()
    
        self.window_stack[-1].close()

    def show_volcano_plot(self):
        if self.data is not None:
              # Assuming 'log2FoldChange' and 'pvalue' columns exist in self.data
            # If not, you may need to calculate these values

            # Calculate -log10(pvalue)
            self.data['-log10(pvalue)'] = -np.log10(self.data['pvalue'])

            # Define significance thresholds
            fc_threshold = 1  # log2 fold change threshold
            p_threshold = 0.05  # p-value threshold

            # Create a new column for significance
            self.data['significant'] = (abs(self.data['log2FoldChange']) > fc_threshold) & (self.data['pvalue'] < p_threshold)

            # Create the volcano plot
            fig = px.scatter(
                self.data, 
                x='log2FoldChange', 
                y='-log10(pvalue)',
                color='significant',
                color_discrete_map={True: 'red', False: 'blue'},
                hover_data=['Gene']  # Assuming 'Gene' is the column with gene names
            )

            # Customize the plot
            fig.update_layout(
                title='Volcano Plot',
                xaxis_title='log2 Fold Change',
                yaxis_title='-log10(p-value)',
                showlegend=True
            )

            # Add threshold lines
            fig.add_hline(y=-np.log10(p_threshold), line_dash="dash", line_color="grey")
            fig.add_vline(x=-fc_threshold, line_dash="dash", line_color="grey")
            fig.add_vline(x=fc_threshold, line_dash="dash", line_color="grey")

            # Show the plot
            fig.show()

            # Save the plot as an HTML file
            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            volcano_path = os.path.join(desktop, "interactive_volcano_plot.html")
            fig.write_html(volcano_path)
            sg.popup(f"Interactive volcano plot saved to {volcano_path}")
        else:
            sg.popup("Please upload data first.")

            fig = px.scatter(self.data, x='log2FoldChange', y='-log10(pvalue)', color='significant')
            fig.show()

    def show_venn_diagram(self):
        file_paths = sg.popup_get_file('Select files (Excel or CSV)', multiple_files=True, file_types=(("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")))
    
        if not file_paths:
          return
    
        file_paths = file_paths.split(';')  # Split the string into a list of file paths
    
        if len(file_paths) < 2:
            sg.popup("At least 2 files are required for a Venn diagram.")
            return
            
        if len(file_paths) < 2:
                sg.popup("At least 2 files are required for a Venn diagram.")
                return

        gene_sets = []
        file_names = []
        for file_path in file_paths:
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path)
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                sg.popup(f"Unsupported file format: {file_path}")
                return
            gene_sets.append(set(df.iloc[:, 0]))
            file_names.append(os.path.basename(file_path))
        
            # Create Venn diagram
        plt.figure(figsize=(10, 10))
        if len(gene_sets) == 2:
            venn2(gene_sets, set_labels=file_names)
        elif len(gene_sets) == 3:
            venn3(gene_sets, set_labels=file_names)
        else:
            sg.popup("Only 2 or 3 files are supported for Venn diagrams.")
            return

        plt.title("Venn Diagram of Gene Sets")

        # Save the diagram to the desktop
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        output_path = os.path.join(desktop_path, 'venn_diagram.png')
        plt.savefig(output_path)
        plt.close()

        # Create an Excel file with unique and common gene names
        unique_genes = [gene_set - set.union(*(gene_sets[:i] + gene_sets[i+1:])) for i, gene_set in enumerate(gene_sets)]
        common_genes = set.intersection(*gene_sets)

        excel_output_path = os.path.join(desktop_path, 'gene_sets.xlsx')
        with pd.ExcelWriter(excel_output_path) as writer:
            for i, gene_set in enumerate(unique_genes):
                pd.DataFrame(list(gene_set), columns=[f'Unique to {file_names[i]}']).to_excel(writer, sheet_name=f'Unique to {file_names[i]}', index=False)
            pd.DataFrame(list(common_genes), columns=['Common Genes']).to_excel(writer, sheet_name='Common Genes', index=False)

        sg.popup(f"Venn diagram saved as 'venn_diagram.png' and gene sets saved as 'gene_sets.xlsx' on your desktop.")

                    


    def run(self):
        sg.theme('Dark Blue 3')  # Set a theme to make the GUI more attractive
        layout = [
            [sg.Text('OmiXInsight', font='Any 20', justification='center')],
            [sg.Button('RNA', size=(20, 2), font='Any 12')],
            [sg.Button('Protein', size=(20, 2), font='Any 12')],
        ]
        self.window_stack.append(sg.Window('Protein Finder', layout, size=(550, 400), element_justification='center'))

        while True:
            event, values = self.window_stack[-1].read()
            if event == 'Protein':
                self.show_protein_options()
            elif event == sg.WINDOW_CLOSED:
                break
        self.window_stack[-1].close()

if __name__ == "__main__":
    app = ProteinFinder()
    app.run()